from __future__ import annotations

import csv
import html
import io
import json
import os
import re
import socket
import sys
import time
import uuid
from datetime import date, timedelta
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib import request
from urllib.parse import parse_qs, urlparse

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - app still supports CSV without openpyxl.
    load_workbook = None


APP_DIR = Path(__file__).resolve().parent
STATIC_DIR = APP_DIR / "static"
DATA_DIR = APP_DIR / "data"
DB_PATH = DATA_DIR / "db.json"
DEFAULT_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4.1-mini")
DEFAULT_HOST = os.environ.get("HOST", "0.0.0.0")


FIELD_ALIASES = {
    "product_name": {"product", "product name", "name", "产品", "产品名", "品名"},
    "model": {"model", "sku", "型号", "货号"},
    "moq": {"moq", "minimum order quantity", "起订量", "最小起订量"},
    "unit_price": {"unit price", "price", "fob price", "单价", "价格"},
    "currency": {"currency", "币种"},
    "packaging": {"packaging", "package", "packing", "包装"},
    "lead_time": {"lead time", "delivery time", "交期", "生产周期"},
    "hs_code": {"hs code", "hscode", "海关编码", "hs编码"},
    "weight": {"weight", "gross weight", "重量", "毛重"},
    "volume": {"volume", "cbm", "体积"},
    "certification": {"certification", "certifications", "certificate", "认证"},
    "tier_prices": {"tier prices", "tier price", "阶梯价", "梯度价格"},
}


def ensure_db() -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not DB_PATH.exists():
        DB_PATH.write_text(json.dumps({"products": [], "inquiries": [], "reminders": []}, ensure_ascii=False, indent=2), encoding="utf-8")
    return json.loads(DB_PATH.read_text(encoding="utf-8"))


def save_db(db: dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")


def access_urls(host: str, port: int) -> list[str]:
    urls = [f"http://127.0.0.1:{port}"]
    if host not in {"0.0.0.0", ""}:
        urls.append(f"http://{host}:{port}")
        return urls
    try:
        addresses = {
            info[4][0]
            for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET)
            if not info[4][0].startswith("127.")
        }
        for address in sorted(addresses):
            urls.append(f"http://{address}:{port}")
    except OSError:
        pass
    return urls


def normalize_header(header: Any) -> str:
    value = str(header or "").strip().lower()
    for canonical, aliases in FIELD_ALIASES.items():
        if value in aliases:
            return canonical
    price_match = re.search(r"(?:price|价格|单价)[_\s-]*(\d+)|(\d+)[_\s-]*(?:pcs|units)?[_\s-]*(?:price|价格|单价)", value)
    if price_match:
        quantity = price_match.group(1) or price_match.group(2)
        return f"price_{quantity}"
    return value.replace(" ", "_")


def normalize_product(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        canonical = normalize_header(key)
        if value is None:
            normalized[canonical] = ""
        else:
            normalized[canonical] = str(value).strip()
    normalized.setdefault("product_name", "")
    normalized.setdefault("model", "")
    normalized.setdefault("moq", "")
    normalized.setdefault("unit_price", "")
    normalized.setdefault("currency", "USD")
    normalized.setdefault("packaging", "")
    normalized.setdefault("lead_time", "")
    normalized.setdefault("hs_code", "")
    normalized.setdefault("weight", "")
    normalized.setdefault("volume", "")
    normalized.setdefault("certification", "")
    normalized.setdefault("tier_prices", "")
    normalized["id"] = normalized.get("id") or str(uuid.uuid4())
    return normalized


def parse_products(filename: str, raw: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [normalize_product(row) for row in reader if any(str(v or "").strip() for v in row.values())]
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ValueError("当前运行环境不支持 XLSX，请上传 CSV。")
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_header(cell) for cell in rows[0]]
        products = []
        for row in rows[1:]:
            item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
            if any(str(v or "").strip() for v in item.values()):
                products.append(normalize_product(item))
        return products
    raise ValueError("请上传 CSV 或 XLSX 产品表。")


def number_from_text(value: Any) -> float | None:
    if value is None:
        return None
    match = re.search(r"\d+(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", ""))


def tokenize(text: str) -> set[str]:
    return {token for token in re.split(r"[^a-zA-Z0-9]+", text.lower()) if len(token) >= 2}


def find_product_matches(inquiry: str, products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    inquiry_tokens = tokenize(inquiry)
    scored = []
    for product in products:
        haystack = " ".join([product.get("product_name", ""), product.get("model", "")])
        product_tokens = tokenize(haystack)
        overlap = inquiry_tokens & product_tokens
        direct_hit = haystack and haystack.lower() in inquiry.lower()
        score = len(overlap) + (5 if direct_hit else 0)
        if score > 0:
            item = dict(product)
            item["match_score"] = score
            scored.append(item)
    return sorted(scored, key=lambda item: item["match_score"], reverse=True)[:3]


def extract_tier_prices(product: dict[str, Any]) -> list[dict[str, Any]]:
    tiers: list[dict[str, Any]] = []
    for key, value in product.items():
        match = re.fullmatch(r"price_(\d+)", str(key))
        if match:
            price = number_from_text(value)
            if price is not None:
                tiers.append({"quantity": int(match.group(1)), "price": price, "source": key})
    raw_tiers = str(product.get("tier_prices") or "")
    for quantity, price in re.findall(r"(\d+)\s*[:=/]\s*(\d+(?:\.\d+)?)", raw_tiers):
        tiers.append({"quantity": int(quantity), "price": float(price), "source": "tier_prices"})
    unique: dict[int, dict[str, Any]] = {}
    for tier in tiers:
        unique[tier["quantity"]] = tier
    return sorted(unique.values(), key=lambda item: item["quantity"])


def select_base_price(product: dict[str, Any], qty: int | None) -> tuple[float | None, str, list[dict[str, Any]]]:
    tiers = extract_tier_prices(product)
    if qty and tiers:
        eligible = [tier for tier in tiers if tier["quantity"] <= qty]
        if eligible:
            tier = eligible[-1]
            return tier["price"], f"{tier['quantity']}+ tier", tiers
    base_price = number_from_text(product.get("unit_price"))
    return base_price, "unit_price", tiers


def extract_inquiry(inquiry: str, products: list[dict[str, Any]]) -> dict[str, Any]:
    email_match = re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", inquiry)
    quantity_match = re.search(r"(\d+(?:,\d{3})*)\s*(pcs|pieces|units|sets|ctns|cartons|kg|tons?|pairs?)", inquiry, re.I)
    incoterm_match = re.search(r"\b(EXW|FOB|FCA|CFR|CIF|CPT|CIP|DAP|DPU|DDP)\b", inquiry, re.I)
    payment_match = re.search(r"\b(T/T|TT|L/C|LC|PayPal|Western Union|OA|D/P|D/A)\b", inquiry, re.I)
    country_candidates = [
        "United States", "USA", "Canada", "Mexico", "Brazil", "Germany", "France", "Italy", "Spain",
        "United Kingdom", "UK", "Australia", "India", "Vietnam", "Thailand", "Indonesia", "Malaysia",
        "Philippines", "UAE", "Saudi Arabia", "South Africa", "Nigeria", "Turkey", "Poland",
    ]
    country = next((country for country in country_candidates if re.search(rf"\b{re.escape(country)}\b", inquiry, re.I)), "")
    company_match = re.search(r"(?:company|company name|from)\s*[:：]\s*(.+)", inquiry, re.I)
    matches = find_product_matches(inquiry, products)
    product_name = matches[0].get("product_name", "") if matches else ""
    missing = []
    if not product_name:
        missing.append("产品名称/型号")
    if not quantity_match:
        missing.append("采购数量")
    if not country:
        missing.append("目的国家")
    if not incoterm_match:
        missing.append("贸易术语")
    if not payment_match:
        missing.append("付款方式")
    if "spec" not in inquiry.lower() and "size" not in inquiry.lower() and "规格" not in inquiry:
        missing.append("关键规格")
    return {
        "customer": {
            "company": company_match.group(1).strip() if company_match else "",
            "email": email_match.group(0) if email_match else "",
            "country": country,
        },
        "request": {
            "product": product_name,
            "quantity": quantity_match.group(0) if quantity_match else "",
            "quantity_value": int(quantity_match.group(1).replace(",", "")) if quantity_match else None,
            "specifications": "",
            "delivery_time": "",
            "incoterm": incoterm_match.group(1).upper() if incoterm_match else "",
            "payment_terms": payment_match.group(1).upper() if payment_match else "",
        },
        "missing_info": missing,
        "matched_products": matches,
    }


def score_inquiry(extracted: dict[str, Any]) -> dict[str, Any]:
    missing_count = len(extracted["missing_info"])
    has_email = bool(extracted["customer"].get("email"))
    has_match = bool(extracted.get("matched_products"))
    qty = extracted["request"].get("quantity_value") or 0
    if not has_email and missing_count >= 4:
        grade = "垃圾询盘"
        score = 18
        reason = "缺少联系方式且关键信息过少。"
    elif has_match and missing_count <= 2 and qty:
        grade = "A"
        score = 88
        reason = "产品、数量和客户信息较完整，可优先报价。"
    elif has_match and missing_count <= 4:
        grade = "B"
        score = 68
        reason = "产品意向明确，但仍需补充部分报价信息。"
    else:
        grade = "C"
        score = 45
        reason = "询盘信息不完整或产品匹配不明确。"
    return {"grade": grade, "score": score, "reason": reason}


def build_quote(extracted: dict[str, Any], settings: dict[str, Any]) -> dict[str, Any]:
    matches = extracted.get("matched_products", [])
    product = matches[0] if matches else None
    margin = number_from_text(settings.get("margin")) or 0
    incoterm = extracted["request"].get("incoterm") or settings.get("incoterm") or "FOB"
    payment = extracted["request"].get("payment_terms") or settings.get("payment") or "30% T/T deposit, 70% before shipment"
    validity = settings.get("validity") or "15 days"
    qty = extracted["request"].get("quantity_value")
    if not product:
        return {
            "available": False,
            "warning": "未找到匹配产品，系统不会编造价格。请补充产品型号或更新产品表。",
            "pi_fields": {},
            "line_items": [],
        }
    base_price, price_source, tiers = select_base_price(product, qty)
    moq = number_from_text(product.get("moq"))
    currency = product.get("currency") or "USD"
    warning = ""
    unit_price = None
    subtotal = None
    if base_price is not None:
        unit_price = round(base_price * (1 + margin / 100), 4)
        subtotal = round(unit_price * qty, 2) if qty else None
    else:
        warning = "产品表缺少单价，请人工补充报价。"
    if moq and qty and qty < moq:
        warning = f"客户数量低于 MOQ（客户 {qty:g}，MOQ {moq:g}），建议先确认是否接受试单加价或提高数量。"
    return {
        "available": True,
        "warning": warning,
        "pi_fields": {
            "seller": "Your Company",
            "buyer": extracted["customer"].get("company") or "To be confirmed",
            "incoterm": incoterm,
            "payment_terms": payment,
            "validity": validity,
            "lead_time": product.get("lead_time") or "To be confirmed",
            "packaging": product.get("packaging") or "To be confirmed",
        },
        "line_items": [
            {
                "product_name": product.get("product_name"),
                "model": product.get("model"),
                "quantity": qty or "To be confirmed",
                "moq": product.get("moq") or "To be confirmed",
                "unit_price": unit_price if unit_price is not None else "To be confirmed",
                "currency": currency,
                "subtotal": subtotal if subtotal is not None else "To be confirmed",
                "price_source": price_source,
            }
        ],
        "pricing_tiers": [
            {
                "quantity": tier["quantity"],
                "price": round(tier["price"] * (1 + margin / 100), 4),
                "currency": currency,
            }
            for tier in tiers
        ],
        "product_notes": {
            "hs_code": product.get("hs_code") or "",
            "weight": product.get("weight") or "",
            "volume": product.get("volume") or "",
            "certification": product.get("certification") or "",
        },
    }


def build_email(extracted: dict[str, Any], score: dict[str, Any], quote: dict[str, Any]) -> str:
    name_line = "Dear Customer,"
    product = extracted["request"].get("product") or "the product you requested"
    if score["grade"] == "垃圾询盘":
        return (
            f"{name_line}\n\n"
            "Thank you for your message. To help us support your request, please share your company name, "
            "target product, quantity, destination country, and any key specifications.\n\n"
            "Best regards,\nSales Team"
        )
    if extracted["missing_info"]:
        questions = "\n".join([f"- {item}" for item in extracted["missing_info"]])
        return (
            f"{name_line}\n\n"
            f"Thank you for your inquiry about {product}. To prepare an accurate quotation, could you please confirm the following details?\n"
            f"{questions}\n\n"
            "Once confirmed, we will send you the quotation and lead time shortly.\n\n"
            "Best regards,\nSales Team"
        )
    line = quote.get("line_items", [{}])[0]
    unit_price = line.get("unit_price", "To be confirmed")
    currency = line.get("currency", "USD")
    return (
        f"{name_line}\n\n"
        f"Thank you for your inquiry about {product}. Based on the information provided, please find our draft quotation below:\n\n"
        f"Product: {line.get('product_name', product)}\n"
        f"Model: {line.get('model', 'To be confirmed')}\n"
        f"Quantity: {line.get('quantity', 'To be confirmed')}\n"
        f"Unit Price: {currency} {unit_price}\n"
        f"Trade Term: {quote.get('pi_fields', {}).get('incoterm', 'FOB')}\n"
        f"Payment Terms: {quote.get('pi_fields', {}).get('payment_terms', 'To be confirmed')}\n"
        f"Lead Time: {quote.get('pi_fields', {}).get('lead_time', 'To be confirmed')}\n\n"
        "This is a draft quotation for your review. Final price and availability will be confirmed after checking specifications and production schedule.\n\n"
        "Best regards,\nSales Team"
    )


def build_followup(extracted: dict[str, Any], score: dict[str, Any], quote: dict[str, Any]) -> dict[str, Any]:
    if score["grade"] == "垃圾询盘":
        stage = "暂停"
        days = 0
        message = "Low-quality inquiry. Keep on record; do not prioritize unless the customer provides more details."
    elif extracted["missing_info"]:
        stage = "待补充信息"
        days = 2
        message = "Just checking whether you had a chance to confirm the missing details so we can prepare an accurate quotation."
    else:
        stage = "已报价待跟进" if quote.get("available") else "待报价"
        days = 3
        message = "Following up on the quotation draft. Please let us know if the price, quantity, or delivery time needs adjustment."
    due_date = "" if days == 0 else (date.today() + timedelta(days=days)).isoformat()
    return {"stage": stage, "due_date": due_date, "message": message}


def fallback_agent(inquiry: str, products: list[dict[str, Any]], settings: dict[str, Any]) -> dict[str, Any]:
    extracted = extract_inquiry(inquiry, products)
    score = score_inquiry(extracted)
    quote = build_quote(extracted, settings)
    email = build_email(extracted, score, quote)
    followup = build_followup(extracted, score, quote)
    return {
        "analysis": extracted,
        "quality": score,
        "quote": quote,
        "email_draft": email,
        "followup": followup,
        "source": "规则引擎兜底",
        "safety": [
            "报价、PI 和邮件均为草稿。",
            "系统不会自动发送邮件。",
            "价格、库存、合规、合同条款需人工最终确认。",
        ],
    }


def call_openai_agent(inquiry: str, products: list[dict[str, Any]], settings: dict[str, Any]) -> dict[str, Any] | None:
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return None
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "analysis": {"type": "object"},
            "quality": {"type": "object"},
            "quote": {"type": "object"},
            "email_draft": {"type": "string"},
            "followup": {"type": "object"},
            "safety": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["analysis", "quality", "quote", "email_draft", "followup", "safety"],
    }
    prompt = {
        "role": "user",
        "content": [
            {
                "type": "input_text",
                "text": (
                    "You are a B2B foreign trade inquiry, quote, and follow-up agent. "
                    "Analyze the inquiry, match only against the supplied products, draft a quotation and English reply. "
                    "Never invent prices or claim final contract status. Return JSON only.\n\n"
                    f"Pricing settings: {json.dumps(settings, ensure_ascii=False)}\n"
                    f"Products: {json.dumps(products[:80], ensure_ascii=False)}\n"
                    f"Inquiry:\n{inquiry}"
                ),
            }
        ],
    }
    payload = {
        "model": DEFAULT_MODEL,
        "input": [prompt],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "trade_agent_result",
                "schema": schema,
                "strict": False,
            }
        },
    }
    try:
        req = request.Request(
            "https://api.openai.com/v1/responses",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            method="POST",
        )
        with request.urlopen(req, timeout=40) as response:
            data = json.loads(response.read().decode("utf-8"))
        text = ""
        for item in data.get("output", []):
            for content in item.get("content", []):
                if content.get("type") in {"output_text", "text"}:
                    text += content.get("text", "")
        if not text and "output_text" in data:
            text = data["output_text"]
        result = json.loads(text)
        result["source"] = f"OpenAI Responses API ({DEFAULT_MODEL})"
        return result
    except Exception as exc:
        print(f"OpenAI call failed, using fallback: {exc}", file=sys.stderr)
        return None


def render_quote_export(record: dict[str, Any]) -> str:
    result = record.get("result") or {}
    analysis = result.get("analysis") or {}
    customer = analysis.get("customer") or {}
    quote = result.get("quote") or {}
    fields = quote.get("pi_fields") or {}
    lines = quote.get("line_items") or []
    tiers = quote.get("pricing_tiers") or []
    notes = quote.get("product_notes") or {}

    def esc(value: Any) -> str:
        return html.escape(str(value if value not in (None, "") else "-"))

    rows = "".join(
        f"<tr><td>{esc(line.get('product_name'))}</td><td>{esc(line.get('model'))}</td>"
        f"<td>{esc(line.get('quantity'))}</td><td>{esc(line.get('currency', 'USD'))} {esc(line.get('unit_price'))}</td>"
        f"<td>{esc(line.get('currency', 'USD'))} {esc(line.get('subtotal'))}</td></tr>"
        for line in lines
    )
    tier_rows = "".join(
        f"<tr><td>{esc(tier.get('quantity'))}+ pcs</td><td>{esc(tier.get('currency', 'USD'))} {esc(tier.get('price'))}</td></tr>"
        for tier in tiers
    )
    note_rows = "".join(
        f"<tr><td>{esc(label)}</td><td>{esc(value)}</td></tr>"
        for label, value in {
            "HS Code": notes.get("hs_code"),
            "Weight": notes.get("weight"),
            "Volume": notes.get("volume"),
            "Certification": notes.get("certification"),
        }.items()
        if value
    )
    warning = quote.get("warning")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Draft Quotation {esc(record.get("id"))}</title>
  <style>
    body {{ font-family: Arial, sans-serif; color: #172026; margin: 36px; }}
    h1 {{ font-size: 26px; margin: 0 0 6px; }}
    .muted {{ color: #63717a; }}
    .banner {{ border: 1px solid #f2c078; background: #fff7ed; padding: 10px; margin: 18px 0; }}
    .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; margin: 18px 0; }}
    .box {{ border: 1px solid #dce4e8; border-radius: 6px; padding: 12px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
    th, td {{ border-bottom: 1px solid #e8eef1; padding: 9px; text-align: left; vertical-align: top; }}
    th {{ background: #f5f7f8; }}
    .actions {{ margin-bottom: 20px; }}
    button {{ border: 0; border-radius: 6px; background: #0f766e; color: white; padding: 9px 13px; cursor: pointer; }}
    @media print {{ .actions {{ display: none; }} body {{ margin: 18mm; }} }}
  </style>
</head>
<body>
  <div class="actions"><button onclick="window.print()">Print / Save as PDF</button></div>
  <h1>Draft Quotation / Proforma Invoice Draft</h1>
  <div class="muted">Generated at {esc(record.get("created_at"))} | Draft only, subject to manual confirmation.</div>
  <div class="banner">This document is not a final contract. Price, stock, compliance, and shipping terms must be confirmed by a human before sending.</div>
  <div class="grid">
    <div class="box">
      <strong>Buyer</strong><br />
      {esc(fields.get("buyer") or customer.get("company") or customer.get("email"))}<br />
      {esc(customer.get("country"))}
    </div>
    <div class="box">
      <strong>Commercial Terms</strong><br />
      Incoterm: {esc(fields.get("incoterm"))}<br />
      Payment: {esc(fields.get("payment_terms"))}<br />
      Validity: {esc(fields.get("validity"))}
    </div>
  </div>
  {f'<div class="banner">{esc(warning)}</div>' if warning else ''}
  <h2>Line Items</h2>
  <table>
    <thead><tr><th>Product</th><th>Model</th><th>Quantity</th><th>Unit Price</th><th>Subtotal</th></tr></thead>
    <tbody>{rows or '<tr><td colspan="5">No matched product. Manual quote required.</td></tr>'}</tbody>
  </table>
  <div class="grid">
    <div>
      <h2>Production & Packing</h2>
      <table><tbody>
        <tr><td>Lead Time</td><td>{esc(fields.get("lead_time"))}</td></tr>
        <tr><td>Packaging</td><td>{esc(fields.get("packaging"))}</td></tr>
        {note_rows}
      </tbody></table>
    </div>
    <div>
      <h2>Tier Prices</h2>
      <table><tbody>{tier_rows or '<tr><td colspan="2">No tier pricing supplied.</td></tr>'}</tbody></table>
    </div>
  </div>
</body>
</html>"""


class Handler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        print(f"[{time.strftime('%H:%M:%S')}] {format % args}")

    def send_json(self, payload: Any, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, content: str, status: int = 200) -> None:
        body = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8") or "{}")

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/" or parsed.path == "":
            self.path = "/static/index.html"
        if parsed.path == "/api/state":
            db = ensure_db()
            self.send_json({
                "products": db["products"],
                "inquiries": db["inquiries"][-20:],
                "reminders": db["reminders"][-30:],
                "openai_enabled": bool(os.environ.get("OPENAI_API_KEY")),
            })
            return
        if parsed.path == "/api/export":
            self.handle_export(parse_qs(parsed.query))
            return
        return super().do_GET()

    def do_POST(self) -> None:
        try:
            if self.path == "/api/upload-products":
                self.handle_upload_products()
                return
            if self.path == "/api/analyze":
                self.handle_analyze()
                return
            if self.path == "/api/stage":
                self.handle_stage_update()
                return
            if self.path == "/api/reset":
                save_db({"products": [], "inquiries": [], "reminders": []})
                self.send_json({"ok": True})
                return
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_upload_products(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        match = re.search(r"boundary=(.+)", content_type)
        if not match:
            raise ValueError("缺少上传文件。")
        boundary = match.group(1).encode("utf-8")
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        parts = body.split(b"--" + boundary)
        filename = "products.csv"
        file_content = b""
        for part in parts:
            if b'name="file"' in part:
                header, _, content = part.partition(b"\r\n\r\n")
                name_match = re.search(rb'filename="([^"]+)"', header)
                if name_match:
                    filename = name_match.group(1).decode("utf-8", errors="ignore")
                file_content = content.rsplit(b"\r\n", 1)[0]
                break
        if not file_content:
            raise ValueError("上传文件为空。")
        products = parse_products(filename, file_content)
        if not products:
            raise ValueError("未从产品表读取到有效产品。")
        db = ensure_db()
        db["products"] = products
        save_db(db)
        self.send_json({"ok": True, "count": len(products), "products": products})

    def handle_analyze(self) -> None:
        payload = self.read_json()
        inquiry = str(payload.get("inquiry", "")).strip()
        if not inquiry:
            raise ValueError("请先粘贴询盘内容。")
        settings = payload.get("settings") or {}
        db = ensure_db()
        result = call_openai_agent(inquiry, db["products"], settings) or fallback_agent(inquiry, db["products"], settings)
        record = {
            "id": str(uuid.uuid4()),
            "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "inquiry": inquiry,
            "result": result,
            "stage": result.get("followup", {}).get("stage", "新询盘"),
        }
        db["inquiries"].append(record)
        followup = dict(result.get("followup") or {})
        if followup.get("due_date"):
            db["reminders"].append({
                "id": str(uuid.uuid4()),
                "created_at": record["created_at"],
                "inquiry_id": record["id"],
                "stage": followup.get("stage", "待跟进"),
                "due_date": followup["due_date"],
                "message": followup.get("message", ""),
                "customer": result.get("analysis", {}).get("customer", {}),
            })
        save_db(db)
        self.send_json({"ok": True, "record": record, "products_count": len(db["products"])})

    def handle_stage_update(self) -> None:
        payload = self.read_json()
        inquiry_id = str(payload.get("inquiry_id") or "")
        reminder_id = str(payload.get("reminder_id") or "")
        stage = str(payload.get("stage") or "").strip()
        if not stage:
            raise ValueError("请选择客户阶段。")
        db = ensure_db()
        updated = False
        for record in db["inquiries"]:
            if record.get("id") == inquiry_id:
                record["stage"] = stage
                record.setdefault("result", {}).setdefault("followup", {})["stage"] = stage
                updated = True
        for reminder in db["reminders"]:
            if reminder.get("id") == reminder_id or reminder.get("inquiry_id") == inquiry_id:
                reminder["stage"] = stage
                if payload.get("due_date") is not None:
                    reminder["due_date"] = str(payload.get("due_date") or "")
                if payload.get("message") is not None:
                    reminder["message"] = str(payload.get("message") or "")
                updated = True
        if not updated:
            raise ValueError("未找到要更新的询盘或跟进任务。")
        save_db(db)
        self.send_json({"ok": True, "stage": stage})

    def handle_export(self, query: dict[str, list[str]]) -> None:
        inquiry_id = (query.get("inquiry_id") or [""])[0]
        db = ensure_db()
        record = next((item for item in db["inquiries"] if item.get("id") == inquiry_id), None)
        if not record:
            self.send_html("<h1>未找到报价记录</h1>", HTTPStatus.NOT_FOUND)
            return
        self.send_html(render_quote_export(record))


def main() -> None:
    ensure_db()
    os.chdir(APP_DIR)
    port = int(os.environ.get("PORT", "8765"))
    host = DEFAULT_HOST
    server = ThreadingHTTPServer((host, port), Handler)
    print("Trade Agent v1.1 running.")
    for url in access_urls(host, port):
        print(f"Open: {url}")
    server.serve_forever()


if __name__ == "__main__":
    main()
